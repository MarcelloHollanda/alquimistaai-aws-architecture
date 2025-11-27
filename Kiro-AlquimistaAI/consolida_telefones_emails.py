"""
Script de Consolidação de Leads - Telefones e Emails

Este script consolida dados de múltiplas planilhas Excel em um único arquivo
formatado para consumo pelo Micro Agente de Disparos & Agendamentos.

AUDITORIA:
- Este arquivo está sujeito a formatação automática pelo Kiro IDE (Autofix)
- Alterações de formatação (espaçamento, indentação) são automáticas e não afetam a lógica
- Para mais informações sobre auditoria e rastreabilidade, consulte:
  docs/prompts/auditoria/LEADS_CONSOLIDADOS_TELEFONES_EMAILS_AUDITORIA.md

FLUXO:
  Planilhas de Origem → [ESTE SCRIPT] → Leads_Consolidados_Telefones_Emails_DEDUP.xlsx
  → Organizador de Leads → Leads_Organizados.xlsx → Micro Agente

Última atualização: 2024-11-27
"""

import pandas as pd
import numpy as np
import re
from pathlib import Path

BASE_DIR = Path(__file__).parent

ARQUIVOS = [
    BASE_DIR / "Planilha Telefones 1.xlsx.xlsx",
    BASE_DIR / "Planilha Telefones 2.xlsx.xlsx",
    BASE_DIR / "Planilha Telefones.xlsx.xlsx",
]

def carregar_aba(nome_aba: str) -> pd.DataFrame:
    frames = []
    for arq in ARQUIVOS:
        if not arq.exists():
            continue
        try:
            df = pd.read_excel(arq, sheet_name=nome_aba)
            df["__source_file"] = arq.name
            frames.append(df)
        except Exception:
            continue
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)

def only_digits(s):
    if pd.isna(s):
        return ""
    return re.sub(r"\D", "", str(s))

def norm_empresa(s):
    if pd.isna(s):
        return ""
    return str(s).strip().lower()

def formatar_cnpj(cnpj_str):
    """Formata CNPJ no padrão 99.999.999/9999-99"""
    if pd.isna(cnpj_str):
        return ""
    
    digits = only_digits(cnpj_str)
    
    # Validar se tem 14 dígitos (CNPJ) ou 11 (CPF)
    if len(digits) == 14:
        # CNPJ: 99.999.999/9999-99
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"
    elif len(digits) == 11:
        # CPF: 999.999.999-99
        return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
    elif len(digits) > 0:
        # Retornar sem formatação se não for CNPJ nem CPF válido
        return digits
    
    return ""

def formatar_telefone(tel_str):
    """Formata telefone no padrão (99)99999-9999 ou (99)9999-9999"""
    if pd.isna(tel_str):
        return ""
    
    digits = only_digits(tel_str)
    
    # Filtrar telefones inválidos (sequências repetidas como 1111111111)
    if len(set(digits)) == 1:  # Todos os dígitos são iguais
        return ""
    
    # Validar tamanho
    if len(digits) == 11:
        # Celular: (99)99999-9999
        return f"({digits[:2]}){digits[2:7]}-{digits[7:]}"
    elif len(digits) == 10:
        # Fixo: (99)9999-9999
        return f"({digits[:2]}){digits[2:6]}-{digits[6:]}"
    elif len(digits) == 13 and digits.startswith("55"):
        # Com código do país +55
        digits = digits[2:]  # Remove o 55
        if len(digits) == 11:
            return f"({digits[:2]}){digits[2:7]}-{digits[7:]}"
        elif len(digits) == 10:
            return f"({digits[:2]}){digits[2:6]}-{digits[6:]}"
    
    # Se não se encaixar nos padrões, retornar vazio
    return ""

EMAIL_REGEX = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")

def explodir_emails(df_emails: pd.DataFrame) -> pd.DataFrame:
    df = df_emails.rename(columns=str.lower).copy()
    if not {"cpf/cnpj", "empresa", "email"}.issubset(df.columns):
        raise ValueError("Aba Emails precisa das colunas: CPF/CNPJ, EMPRESA, EMAIL")
    
    linhas = []
    for _, row in df.iterrows():
        doc = row["cpf/cnpj"]
        emp = row["empresa"]
        raw = row["email"]
        
        if pd.isna(raw) or str(raw).strip() == "":
            linhas.append({"cpf/cnpj": doc, "empresa": emp, "email": None})
            continue
        
        text = str(raw)
        encontrados = EMAIL_REGEX.findall(text)
        if encontrados:
            for e in encontrados:
                linhas.append({"cpf/cnpj": doc, "empresa": emp, "email": e.strip()})
        else:
            # preserva bruto, mesmo que estranho
            linhas.append({"cpf/cnpj": doc, "empresa": emp, "email": text.strip()})
    
    return pd.DataFrame(linhas)

def explodir_telefones(df_tel: pd.DataFrame) -> pd.DataFrame:
    df = df_tel.rename(columns=str.lower).copy()
    if not {"cpf/cnpj", "empresa", "telefone"}.issubset(df.columns):
        raise ValueError("Aba Telefones precisa das colunas: CPF/CNPJ, EMPRESA, TELEFONE")
    
    linhas = []
    for _, row in df.iterrows():
        doc = row["cpf/cnpj"]
        emp = row["empresa"]
        raw = row["telefone"]
        
        if pd.isna(raw) or str(raw).strip() == "":
            linhas.append({"cpf/cnpj": doc, "empresa": emp, "telefone": None})
            continue
        
        text = str(raw)
        # separa por delimitadores comuns
        partes = re.split(r"[;|\n,/]", text)
        encontrou = False
        for p in partes:
            p = p.strip()
            if not p:
                continue
            encontrou = True
            linhas.append({"cpf/cnpj": doc, "empresa": emp, "telefone": p})
        if not encontrou:
            linhas.append({"cpf/cnpj": doc, "empresa": emp, "telefone": text.strip()})
    
    return pd.DataFrame(linhas)

def construir_chave_entidade(df: pd.DataFrame) -> pd.DataFrame:
    df["__doc_norm"] = df["cpf/cnpj"].apply(only_digits)
    df["__emp_norm"] = df["empresa"].apply(norm_empresa)
    
    def make_key(row):
        if row["__doc_norm"]:
            return "D:" + row["__doc_norm"]
        if row["__emp_norm"]:
            return "E:" + row["__emp_norm"]
        return "Z:sem_chave"
    
    df["__entity"] = df.apply(make_key, axis=1)
    return df

def consolidar_modelo_A():
    # 1) carregar abas
    tel_raw = carregar_aba("Telefones")
    mail_raw = carregar_aba("Emails")
    
    if tel_raw.empty and mail_raw.empty:
        raise RuntimeError("Não encontrei abas Telefones/Emails nos arquivos.")
    
    # 2) explodir
    tel_exp = explodir_telefones(tel_raw)
    mail_exp = explodir_emails(mail_raw)
    
    # 3) construir chave entidade
    tel_exp = construir_chave_entidade(tel_exp)
    mail_exp = construir_chave_entidade(mail_exp)
    
    # 4) agrupar telefones/emails por entidade
    tel_group = tel_exp.groupby("__entity")["telefone"].apply(list)
    mail_group = mail_exp.groupby("__entity")["email"].apply(list)
    
    # representantes
    doc_rep_tel = tel_exp.groupby("__entity")["cpf/cnpj"].first()
    emp_rep_tel = tel_exp.groupby("__entity")["empresa"].first()
    doc_rep_mail = mail_exp.groupby("__entity")["cpf/cnpj"].first()
    emp_rep_mail = mail_exp.groupby("__entity")["empresa"].first()
    
    all_keys = set(tel_group.index).union(set(mail_group.index))
    
    rows = []
    for key in all_keys:
        if key in doc_rep_tel.index:
            doc_val = doc_rep_tel.loc[key]
            emp_val = emp_rep_tel.loc[key]
        else:
            doc_val = doc_rep_mail.loc[key] if key in doc_rep_mail.index else None
            emp_val = emp_rep_mail.loc[key] if key in emp_rep_mail.index else None
        
        tels = list(tel_group.get(key, []))
        mails = list(mail_group.get(key, []))
        
        if tels and mails:
            for t in tels:
                for m in mails:
                    rows.append({
                        "CNPJ/CPF": doc_val,
                        "Empresa": emp_val,
                        "Telefone": t,
                        "Email": m,
                        "__entity": key,
                    })
        elif tels:
            for t in tels:
                rows.append({
                    "CNPJ/CPF": doc_val,
                    "Empresa": emp_val,
                    "Telefone": t,
                    "Email": None,
                    "__entity": key,
                })
        elif mails:
            for m in mails:
                rows.append({
                    "CNPJ/CPF": doc_val,
                    "Empresa": emp_val,
                    "Telefone": None,
                    "Email": m,
                    "__entity": key,
                })
    
    cons = pd.DataFrame(rows)
    return cons

def deduplicar_por_entidade(cons: pd.DataFrame) -> pd.DataFrame:
    df = cons.copy()
    
    # Formatar CNPJ/CPF e Telefone
    df["CNPJ/CPF_formatado"] = df["CNPJ/CPF"].apply(formatar_cnpj)
    df["Telefone_formatado"] = df["Telefone"].apply(formatar_telefone)
    
    # Filtrar registros com telefone inválido (vazio após formatação)
    df = df[df["Telefone_formatado"] != ""]
    
    df["__tel_norm"] = df["Telefone_formatado"].fillna("").astype(str).str.strip()
    df["__mail_norm"] = df["Email"].fillna("").astype(str).str.strip().str.lower()
    
    df_dedup = df.drop_duplicates(subset=["__entity", "__tel_norm", "__mail_norm"])
    
    # montar saída final
    saida = pd.DataFrame()
    saida["Empresa"] = df_dedup["Empresa"]
    def contato_from_email(e):
        if pd.isna(e) or str(e).strip() == "" or "@" not in str(e):
            return ""
        return str(e).split("@", 1)[0]
    saida["Contato"] = df_dedup["Email"].apply(contato_from_email)
    saida["CNPJ/CPF"] = df_dedup["CNPJ/CPF_formatado"]
    saida["Email"] = df_dedup["Email"]
    saida["Telefone"] = df_dedup["Telefone_formatado"]
    
    return saida

def formatar_excel(df: pd.DataFrame, caminho: Path):
    """Formata o arquivo Excel para melhor visualização"""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_TEXT

    # Converter CNPJ/CPF e Telefone para string com prefixo para evitar notação científica
    df_copy = df.copy()
    if "CNPJ/CPF" in df_copy.columns:
        df_copy["CNPJ/CPF"] = df_copy["CNPJ/CPF"].apply(lambda x: str(x) if pd.notna(x) else "")
    if "Telefone" in df_copy.columns:
        df_copy["Telefone"] = df_copy["Telefone"].apply(lambda x: str(x) if pd.notna(x) else "")

    # Salvar DataFrame básico
    df_copy.to_excel(caminho, index=False)

    # Carregar workbook para formatação
    wb = load_workbook(caminho)
    ws = wb.active

    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Aplicar estilo ao cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Identificar colunas de CNPJ/CPF e Telefone
    cnpj_col = None
    telefone_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "CNPJ/CPF":
            cnpj_col = idx
        elif cell.value == "Telefone":
            telefone_col = idx

    # Forçar formato de texto para CNPJ/CPF e Telefone
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx, cell in enumerate(row, 1):
            # Aplicar formato de texto para CNPJ e Telefone
            if idx == cnpj_col or idx == telefone_col:
                cell.number_format = FORMAT_TEXT
                # Garantir que o valor seja string
                if cell.value is not None:
                    cell.value = str(cell.value)
            
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
        ws.column_dimensions[column_letter].width = adjusted_width

    # Adicionar filtros automáticos
    ws.auto_filter.ref = ws.dimensions

    # Congelar primeira linha
    ws.freeze_panes = "A2"

    # Salvar workbook formatado
    wb.save(caminho)


def main():
    cons = consolidar_modelo_A()
    final = deduplicar_por_entidade(cons)
    
    out_xlsx = BASE_DIR / "Leads_Consolidados_Telefones_Emails_DEDUP.xlsx"
    out_csv  = BASE_DIR / "Leads_Consolidados_Telefones_Emails_DEDUP.csv"
    
    # Gerar Excel formatado
    formatar_excel(final, out_xlsx)
    
    # Gerar CSV simples
    final.to_csv(out_csv, index=False)
    
    print("Registros consolidados:", len(final))
    print("Arquivos gerados:")
    print(" -", out_xlsx, "(formatado)")
    print(" -", out_csv)

if __name__ == "__main__":
    main()
